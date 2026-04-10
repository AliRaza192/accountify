"""
BI & Analytics Dashboard Service
Provides KPI metrics, trends, financial ratios, and export functionality.
"""

import logging
import io
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Dict, Any, Optional
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


class BIDashboardService:
    """Service for generating BI dashboard metrics and analytics."""

    def __init__(self, db: Session):
        self.db = db

    # ==================== KPI METRICS ====================

    def get_kpi_metrics(
        self,
        company_id: UUID,
        start_date: date,
        end_date: date,
    ) -> Dict[str, Decimal]:
        """
        Returns KPI cards data:
        - Total Revenue, Total Expenses, Net Profit
        - Gross Margin %, Net Profit %
        - Current Ratio, Quick Ratio, DSO
        """
        total_revenue = self._get_total_revenue(company_id, start_date, end_date)
        total_expenses = self._get_total_expenses(company_id, start_date, end_date)
        cogs = self._get_cogs(company_id, start_date, end_date)
        gross_profit = total_revenue - cogs
        net_profit = total_revenue - total_expenses

        gross_margin_pct = self._safe_percent(gross_profit, total_revenue) * Decimal("100")
        net_profit_pct = self._safe_percent(net_profit, total_revenue) * Decimal("100")

        current_ratio, quick_ratio = self._get_liquidity_ratios(company_id)
        dso = self._get_dso(company_id, start_date, end_date, total_revenue)

        return {
            "total_revenue": total_revenue,
            "total_expenses": total_expenses,
            "net_profit": net_profit,
            "gross_margin_percent": gross_margin_pct,
            "net_profit_percent": net_profit_pct,
            "current_ratio": current_ratio,
            "quick_ratio": quick_ratio,
            "dso": dso,
        }

    # ==================== REVENUE TRENDS ====================

    def get_revenue_trends(
        self,
        company_id: UUID,
        months: int = 12,
    ) -> List[Dict[str, Any]]:
        """
        Monthly revenue data with YoY comparison.
        Returns list of {month, value, previous_year}.
        """
        end_date = date.today()
        start_date = end_date - timedelta(days=months * 30)

        # Current period monthly revenue
        current_query = text("""
            SELECT
                TO_CHAR(invoice_date, 'YYYY-MM') AS month,
                COALESCE(SUM(total_amount), 0) AS value
            FROM invoices
            WHERE company_id = :company_id
              AND invoice_date >= :start_date
              AND invoice_date <= :end_date
              AND is_deleted = false
              AND status IN ('paid', 'sent', 'overdue')
            GROUP BY TO_CHAR(invoice_date, 'YYYY-MM')
            ORDER BY month
        """)
        current_result = self.db.execute(
            current_query,
            {"company_id": company_id, "start_date": start_date, "end_date": end_date},
        )
        current_map = {row.month: float(row.value) for row in current_result}

        # Previous year same months
        prev_start = start_date.replace(year=start_date.year - 1)
        prev_end = end_date.replace(year=end_date.year - 1)
        prev_query = text("""
            SELECT
                TO_CHAR(invoice_date, 'YYYY-MM') AS month,
                COALESCE(SUM(total_amount), 0) AS value
            FROM invoices
            WHERE company_id = :company_id
              AND invoice_date >= :start_date
              AND invoice_date <= :end_date
              AND is_deleted = false
              AND status IN ('paid', 'sent', 'overdue')
            GROUP BY TO_CHAR(invoice_date, 'YYYY-MM')
            ORDER BY month
        """)
        prev_result = self.db.execute(
            prev_query,
            {"company_id": company_id, "start_date": prev_start, "end_date": prev_end},
        )
        prev_map = {row.month: float(row.value) for row in prev_result}

        # Build 12-month series
        trends: List[Dict[str, Any]] = []
        for i in range(months - 1, -1, -1):
            dt = end_date - timedelta(days=i * 30)
            month_key = dt.strftime("%Y-%m")
            # Use actual month boundaries
            if dt.month == 12:
                next_month = dt.replace(year=dt.year + 1, month=1, day=1)
            else:
                next_month = dt.replace(month=dt.month + 1, day=1)
            month_label = dt.strftime("%Y-%m")
            prev_month_label = f"{dt.year - 1}-{dt.month:02d}"

            trends.append({
                "month": month_label,
                "value": current_map.get(month_label, 0.0),
                "previous_year": prev_map.get(prev_month_label, 0.0),
            })

        return trends

    # ==================== EXPENSE TRENDS ====================

    def get_expense_trends(
        self,
        company_id: UUID,
        months: int = 12,
    ) -> Dict[str, Any]:
        """
        Monthly expense breakdown by category.
        Returns {trends, by_category, total_expenses}.
        """
        end_date = date.today()
        start_date = end_date - timedelta(days=months * 30)

        # Monthly expense totals
        monthly_query = text("""
            SELECT
                TO_CHAR(bill_date, 'YYYY-MM') AS month,
                COALESCE(SUM(total_amount), 0) AS value
            FROM bills
            WHERE company_id = :company_id
              AND bill_date >= :start_date
              AND bill_date <= :end_date
              AND is_deleted = false
            GROUP BY TO_CHAR(bill_date, 'YYYY-MM')
            ORDER BY month
        """)
        monthly_result = self.db.execute(
            monthly_query,
            {"company_id": company_id, "start_date": start_date, "end_date": end_date},
        )
        monthly_map = {row.month: float(row.value) for row in monthly_result}

        # Build trend series
        trends: List[Dict[str, Any]] = []
        for i in range(months - 1, -1, -1):
            dt = end_date - timedelta(days=i * 30)
            month_label = dt.strftime("%Y-%m")
            prev_label = f"{dt.year - 1}-{dt.month:02d}"

            # Previous year
            prev_query = text("""
                SELECT COALESCE(SUM(total_amount), 0) AS value
                FROM bills
                WHERE company_id = :company_id
                  AND TO_CHAR(bill_date, 'YYYY-MM') = :month
                  AND is_deleted = false
            """)
            prev_result = self.db.execute(
                prev_query,
                {"company_id": company_id, "month": prev_label},
            )
            prev_val = float(prev_result.scalar() or 0)

            trends.append({
                "month": month_label,
                "value": monthly_map.get(month_label, 0.0),
                "previous_year": prev_val,
            })

        # Category breakdown (using account types from journal entries linked to bills)
        by_category = self._get_expense_by_category(company_id, start_date, end_date)
        total_expenses = sum(Decimal(str(t["value"])) for t in trends)

        return {
            "trends": trends,
            "by_category": by_category,
            "total_expenses": total_expenses,
        }

    # ==================== FINANCIAL RATIOS ====================

    def get_financial_ratios(
        self,
        company_id: UUID,
        start_date: date,
        end_date: date,
    ) -> Dict[str, Any]:
        """Calculate all financial ratios."""
        total_revenue = self._get_total_revenue(company_id, start_date, end_date)
        total_expenses = self._get_total_expenses(company_id, start_date, end_date)
        cogs = self._get_cogs(company_id, start_date, end_date)
        gross_profit = total_revenue - cogs
        net_profit = total_revenue - total_expenses

        current_ratio, quick_ratio = self._get_liquidity_ratios(company_id)
        cash_ratio = self._get_cash_ratio(company_id)
        dso = self._get_dso(company_id, start_date, end_date, total_revenue)
        dpo = self._get_dpo(company_id, start_date, end_date, total_expenses)

        total_assets = self._get_total_assets(company_id)
        total_liabilities = self._get_total_liabilities(company_id)
        total_equity = total_assets - total_liabilities

        roa = self._safe_percent(net_profit, total_assets) * Decimal("100")
        roe = self._safe_percent(net_profit, total_equity) * Decimal("100")
        gross_margin = self._safe_percent(gross_profit, total_revenue) * Decimal("100")
        net_margin = self._safe_percent(net_profit, total_revenue) * Decimal("100")
        debt_to_equity = self._safe_percent(total_liabilities, total_equity)
        debt_ratio = self._safe_percent(total_liabilities, total_assets)

        return {
            # Liquidity
            "current_ratio": current_ratio,
            "quick_ratio": quick_ratio,
            "cash_ratio": cash_ratio,
            # Profitability
            "gross_profit_margin": gross_margin,
            "net_profit_margin": net_margin,
            "return_on_assets": roa,
            "return_on_equity": roe,
            # Efficiency
            "dso": dso,
            "dpo": dpo,
            # Leverage
            "debt_to_equity": debt_to_equity,
            "debt_ratio": debt_ratio,
            # Totals
            "total_revenue": total_revenue,
            "total_expenses": total_expenses,
            "net_profit": net_profit,
            "total_assets": total_assets,
            "total_liabilities": total_liabilities,
            "total_equity": total_equity,
        }

    # ==================== TOP CUSTOMERS ====================

    def get_top_customers(
        self,
        company_id: UUID,
        limit: int = 10,
    ) -> List[Dict[str, Any]]:
        """Top customers by revenue."""
        query = text("""
            SELECT
                c.id AS customer_id,
                c.name AS customer_name,
                COALESCE(SUM(i.total_amount), 0) AS total_revenue,
                COUNT(i.id) AS invoice_count,
                COALESCE(AVG(i.total_amount), 0) AS avg_invoice_value,
                COALESCE(SUM(i.total_amount - i.paid_amount), 0) AS outstanding_balance
            FROM customers c
            LEFT JOIN invoices i ON i.customer_id = c.id
                AND i.is_deleted = false
                AND i.status IN ('paid', 'sent', 'overdue')
            WHERE c.company_id = :company_id
              AND c.is_deleted = false
            GROUP BY c.id, c.name
            ORDER BY total_revenue DESC
            LIMIT :limit
        """)
        result = self.db.execute(query, {"company_id": company_id, "limit": limit})
        rows = result.fetchall()

        return [
            {
                "customer_id": str(row.customer_id),
                "customer_name": row.customer_name,
                "total_revenue": Decimal(str(row.total_revenue)),
                "invoice_count": row.invoice_count,
                "avg_invoice_value": Decimal(str(row.avg_invoice_value)),
                "outstanding_balance": Decimal(str(row.outstanding_balance)),
            }
            for row in rows
        ]

    # ==================== TOP PRODUCTS ====================

    def get_top_products(
        self,
        company_id: UUID,
        limit: int = 10,
    ) -> List[Dict[str, Any]]:
        """Top products by sales."""
        query = text("""
            SELECT
                p.id AS product_id,
                p.name AS product_name,
                p.sku,
                COALESCE(SUM(ii.total_price), 0) AS total_sales,
                COALESCE(SUM(ii.quantity), 0) AS quantity_sold,
                COALESCE(AVG(ii.unit_price), 0) AS avg_unit_price
            FROM products p
            LEFT JOIN invoice_items ii ON ii.product_id = p.id
            LEFT JOIN invoices i ON i.id = ii.invoice_id
                AND i.is_deleted = false
                AND i.status IN ('paid', 'sent', 'overdue')
            WHERE p.company_id = :company_id
              AND p.is_deleted = false
            GROUP BY p.id, p.name, p.sku
            ORDER BY total_sales DESC
            LIMIT :limit
        """)
        result = self.db.execute(query, {"company_id": company_id, "limit": limit})
        rows = result.fetchall()

        return [
            {
                "product_id": str(row.product_id),
                "product_name": row.product_name,
                "sku": row.sku,
                "total_sales": Decimal(str(row.total_sales)),
                "quantity_sold": int(row.quantity_sold),
                "avg_unit_price": Decimal(str(row.avg_unit_price)),
            }
            for row in rows
        ]

    # ==================== EXPORT TO EXCEL ====================

    def export_to_excel(
        self,
        company_id: UUID,
        start_date: date,
        end_date: date,
    ) -> Dict[str, Any]:
        """
        Generate Excel file with all metrics.
        Returns {success, file_name, file_url, message}.
        Uses openpyxl if available, otherwise returns CSV fallback.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # --- Sheet 1: KPI Summary ---
            ws_kpi = wb.active
            ws_kpi.title = "KPI Summary"

            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            kpi_data = self.get_kpi_metrics(company_id, start_date, end_date)
            ws_kpi.append(["BI Dashboard Export", ""])
            ws_kpi.append([f"Period: {start_date} to {end_date}", ""])
            ws_kpi.append([])
            ws_kpi.append(["Metric", "Value"])
            for cell in ws_kpi[4]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for key, value in kpi_data.items():
                label = key.replace("_", " ").title()
                ws_kpi.append([label, float(value)])

            ws_kpi.column_dimensions["A"].width = 30
            ws_kpi.column_dimensions["B"].width = 20

            # --- Sheet 2: Revenue Trends ---
            ws_rev = wb.create_sheet("Revenue Trends")
            ws_rev.append(["Month", "Revenue", "Previous Year"])
            for cell in ws_rev[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            trends = self.get_revenue_trends(company_id)
            for t in trends:
                ws_rev.append([t["month"], t["value"], t["previous_year"]])

            ws_rev.column_dimensions["A"].width = 15
            ws_rev.column_dimensions["B"].width = 18
            ws_rev.column_dimensions["C"].width = 18

            # --- Sheet 3: Expense Trends ---
            ws_exp = wb.create_sheet("Expense Trends")
            ws_exp.append(["Month", "Expenses", "Previous Year"])
            for cell in ws_exp[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            expense_data = self.get_expense_trends(company_id)
            for t in expense_data["trends"]:
                ws_exp.append([t["month"], t["value"], t["previous_year"]])

            ws_exp.append([])
            ws_exp.append(["Category Breakdown", ""])
            ws_exp.append(["Category", "Amount", "Percentage"])
            for cell in ws_exp[ws_exp.max_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for cat in expense_data["by_category"]:
                ws_exp.append([cat["category"], float(cat["amount"]), float(cat["percentage"])])

            ws_exp.column_dimensions["A"].width = 25
            ws_exp.column_dimensions["B"].width = 18
            ws_exp.column_dimensions["C"].width = 15

            # --- Sheet 4: Financial Ratios ---
            ws_ratios = wb.create_sheet("Financial Ratios")
            ws_ratios.append(["Ratio", "Value"])
            for cell in ws_ratios[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            ratios = self.get_financial_ratios(company_id, start_date, end_date)
            for key, value in ratios.items():
                label = key.replace("_", " ").title()
                ws_ratios.append([label, float(value)])

            ws_ratios.column_dimensions["A"].width = 30
            ws_ratios.column_dimensions["B"].width = 20

            # --- Sheet 5: Top Customers ---
            ws_cust = wb.create_sheet("Top Customers")
            ws_cust.append(["Customer", "Revenue", "Invoices", "Avg Invoice", "Outstanding"])
            for cell in ws_cust[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            customers = self.get_top_customers(company_id)
            for c in customers:
                ws_cust.append([
                    c["customer_name"],
                    float(c["total_revenue"]),
                    c["invoice_count"],
                    float(c["avg_invoice_value"]),
                    float(c["outstanding_balance"]),
                ])

            ws_cust.column_dimensions["A"].width = 30
            ws_cust.column_dimensions["B"].width = 18
            ws_cust.column_dimensions["C"].width = 12
            ws_cust.column_dimensions["D"].width = 18
            ws_cust.column_dimensions["E"].width = 18

            # --- Sheet 6: Top Products ---
            ws_prod = wb.create_sheet("Top Products")
            ws_prod.append(["Product", "SKU", "Total Sales", "Qty Sold", "Avg Price"])
            for cell in ws_prod[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            products = self.get_top_products(company_id)
            for p in products:
                ws_prod.append([
                    p["product_name"],
                    p["sku"] or "",
                    float(p["total_sales"]),
                    p["quantity_sold"],
                    float(p["avg_unit_price"]),
                ])

            ws_prod.column_dimensions["A"].width = 30
            ws_prod.column_dimensions["B"].width = 15
            ws_prod.column_dimensions["C"].width = 18
            ws_prod.column_dimensions["D"].width = 12
            ws_prod.column_dimensions["E"].width = 15

            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            file_name = f"bi_dashboard_{company_id}_{start_date}_{end_date}.xlsx"

            # In production, upload to storage and return URL.
            # For now, return the bytes via base64 or file path.
            import base64
            file_b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

            return {
                "success": True,
                "file_name": file_name,
                "file_url": f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{file_b64}",
                "message": "Excel export generated successfully",
            }

        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self._export_csv_fallback(company_id, start_date, end_date)
        except Exception as e:
            logger.error(f"Export failed: {e}", exc_info=True)
            return {
                "success": False,
                "file_name": None,
                "file_url": None,
                "message": f"Export failed: {str(e)}",
            }

    def _export_csv_fallback(
        self,
        company_id: UUID,
        start_date: date,
        end_date: date,
    ) -> Dict[str, Any]:
        """Fallback CSV export when openpyxl is not available."""
        import csv
        import base64

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # KPI
        kpi = self.get_kpi_metrics(company_id, start_date, end_date)
        writer.writerow(["KPI Summary"])
        writer.writerow(["Metric", "Value"])
        for k, v in kpi.items():
            writer.writerow([k.replace("_", " ").title(), v])
        writer.writerow([])

        # Top Customers
        writer.writerow(["Top Customers"])
        writer.writerow(["Customer", "Revenue", "Invoices", "Avg Invoice", "Outstanding"])
        for c in self.get_top_customers(company_id):
            writer.writerow([
                c["customer_name"], c["total_revenue"],
                c["invoice_count"], c["avg_invoice_value"],
                c["outstanding_balance"],
            ])
        writer.writerow([])

        # Top Products
        writer.writerow(["Top Products"])
        writer.writerow(["Product", "SKU", "Total Sales", "Qty Sold", "Avg Price"])
        for p in self.get_top_products(company_id):
            writer.writerow([
                p["product_name"], p["sku"] or "",
                p["total_sales"], p["quantity_sold"],
                p["avg_unit_price"],
            ])

        file_name = f"bi_dashboard_{company_id}_{start_date}_{end_date}.csv"
        file_b64 = base64.b64encode(buffer.getvalue().encode("utf-8")).decode("utf-8")

        return {
            "success": True,
            "file_name": file_name,
            "file_url": f"data:text/csv;base64,{file_b64}",
            "message": "CSV export generated (openpyxl not available)",
        }

    # ==================== INTERNAL HELPERS ====================

    def _get_total_revenue(
        self, company_id: UUID, start_date: date, end_date: date
    ) -> Decimal:
        """Total revenue from invoices in period."""
        query = text("""
            SELECT COALESCE(SUM(total_amount), 0) AS total
            FROM invoices
            WHERE company_id = :company_id
              AND invoice_date >= :start_date
              AND invoice_date <= :end_date
              AND is_deleted = false
              AND status IN ('paid', 'sent', 'overdue')
        """)
        result = self.db.execute(
            query,
            {"company_id": company_id, "start_date": start_date, "end_date": end_date},
        )
        return Decimal(str(result.scalar() or 0))

    def _get_total_expenses(
        self, company_id: UUID, start_date: date, end_date: date
    ) -> Decimal:
        """Total expenses from bills in period."""
        query = text("""
            SELECT COALESCE(SUM(total_amount), 0) AS total
            FROM bills
            WHERE company_id = :company_id
              AND bill_date >= :start_date
              AND bill_date <= :end_date
              AND is_deleted = false
        """)
        result = self.db.execute(
            query,
            {"company_id": company_id, "start_date": start_date, "end_date": end_date},
        )
        return Decimal(str(result.scalar() or 0))

    def _get_cogs(
        self, company_id: UUID, start_date: date, end_date: date
    ) -> Decimal:
        """
        Cost of Goods Sold: sum of (quantity * cost_price) from invoice_items.
        Falls back to journal entries for COGS accounts if needed.
        """
        query = text("""
            SELECT COALESCE(SUM(ii.quantity * p.cost_price), 0) AS total_cogs
            FROM invoice_items ii
            JOIN invoices i ON i.id = ii.invoice_id
            LEFT JOIN products p ON p.id = ii.product_id
            WHERE i.company_id = :company_id
              AND i.invoice_date >= :start_date
              AND i.invoice_date <= :end_date
              AND i.is_deleted = false
              AND i.status IN ('paid', 'sent', 'overdue')
        """)
        result = self.db.execute(
            query,
            {"company_id": company_id, "start_date": start_date, "end_date": end_date},
        )
        return Decimal(str(result.scalar() or 0))

    def _get_liquidity_ratios(
        self, company_id: UUID
    ) -> tuple[Decimal, Decimal]:
        """
        Current Ratio = Current Assets / Current Liabilities
        Quick Ratio = (Current Assets - Inventory) / Current Liabilities
        Uses account balances from chart of accounts.
        """
        # Get current assets (account types like 'current_asset')
        ca_query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type = 'current_asset'
              AND is_deleted = false
        """)
        ca_result = self.db.execute(ca_query, {"company_id": company_id})
        current_assets = Decimal(str(ca_result.scalar() or 0))

        # Get current liabilities
        cl_query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type = 'current_liability'
              AND is_deleted = false
        """)
        cl_result = self.db.execute(cl_query, {"company_id": company_id})
        current_liabilities = Decimal(str(cl_result.scalar() or 0))

        # Get inventory value
        inv_query = text("""
            SELECT COALESCE(SUM(stock_quantity * cost_price), 0) AS total
            FROM products
            WHERE company_id = :company_id
              AND is_deleted = false
        """)
        inv_result = self.db.execute(inv_query, {"company_id": company_id})
        inventory = Decimal(str(inv_result.scalar() or 0))

        current_ratio = self._safe_percent(current_assets, current_liabilities)
        quick_assets = current_assets - inventory
        quick_ratio = self._safe_percent(quick_assets, current_liabilities)

        return current_ratio, quick_ratio

    def _get_cash_ratio(self, company_id: UUID) -> Decimal:
        """Cash Ratio = Cash & Equivalents / Current Liabilities."""
        cash_query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type = 'cash'
              AND is_deleted = false
        """)
        cash_result = self.db.execute(cash_query, {"company_id": company_id})
        cash = Decimal(str(cash_result.scalar() or 0))

        cl_query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type = 'current_liability'
              AND is_deleted = false
        """)
        cl_result = self.db.execute(cl_query, {"company_id": company_id})
        current_liabilities = Decimal(str(cl_result.scalar() or 0))

        return self._safe_percent(cash, current_liabilities)

    def _get_dso(
        self, company_id: UUID, start_date: date, end_date: date, total_revenue: Decimal
    ) -> Decimal:
        """
        DSO = (Accounts Receivable / Total Revenue) * Number of Days in Period.
        """
        ar_query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type = 'accounts_receivable'
              AND is_deleted = false
        """)
        ar_result = self.db.execute(ar_query, {"company_id": company_id})
        ar = Decimal(str(ar_result.scalar() or 0))

        days = (end_date - start_date).days or 1
        dso = self._safe_percent(ar, total_revenue) * Decimal(str(days))
        return dso.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _get_dpo(
        self, company_id: UUID, start_date: date, end_date: date, total_expenses: Decimal
    ) -> Decimal:
        """
        DPO = (Accounts Payable / Total Expenses) * Number of Days in Period.
        """
        ap_query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type = 'accounts_payable'
              AND is_deleted = false
        """)
        ap_result = self.db.execute(ap_query, {"company_id": company_id})
        ap = Decimal(str(ap_result.scalar() or 0))

        days = (end_date - start_date).days or 1
        dpo = self._safe_percent(ap, total_expenses) * Decimal(str(days))
        return dpo.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _get_total_assets(self, company_id: UUID) -> Decimal:
        """Sum of all asset account balances."""
        query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type IN ('asset', 'current_asset', 'fixed_asset', 'cash')
              AND is_deleted = false
        """)
        result = self.db.execute(query, {"company_id": company_id})
        return Decimal(str(result.scalar() or 0))

    def _get_total_liabilities(self, company_id: UUID) -> Decimal:
        """Sum of all liability account balances."""
        query = text("""
            SELECT COALESCE(SUM(balance), 0) AS total
            FROM accounts
            WHERE company_id = :company_id
              AND type IN ('liability', 'current_liability', 'long_term_liability')
              AND is_deleted = false
        """)
        result = self.db.execute(query, {"company_id": company_id})
        return Decimal(str(result.scalar() or 0))

    def _get_expense_by_category(
        self, company_id: UUID, start_date: date, end_date: date
    ) -> List[Dict[str, Any]]:
        """
        Expense breakdown by account type/category.
        Uses journal entries linked to expense accounts.
        """
        query = text("""
            SELECT
                a.type AS category,
                COALESCE(SUM(je.debit), 0) AS amount
            FROM journal_entries je
            JOIN journals j ON j.id = je.journal_id
            JOIN accounts a ON a.id = je.account_id
            WHERE j.company_id = :company_id
              AND j.entry_date >= :start_date
              AND j.entry_date <= :end_date
              AND j.is_posted = true
              AND a.type IN ('expense', 'cost_of_goods_sold')
              AND a.is_deleted = false
            GROUP BY a.type
            ORDER BY amount DESC
        """)
        result = self.db.execute(
            query,
            {"company_id": company_id, "start_date": start_date, "end_date": end_date},
        )
        rows = result.fetchall()

        if not rows:
            # Fallback: use bill categories (vendor-based)
            bill_query = text("""
                SELECT
                    v.name AS category,
                    COALESCE(SUM(b.total_amount), 0) AS amount
                FROM bills b
                JOIN vendors v ON v.id = b.vendor_id
                WHERE b.company_id = :company_id
                  AND b.bill_date >= :start_date
                  AND b.bill_date <= :end_date
                  AND b.is_deleted = false
                GROUP BY v.id, v.name
                ORDER BY amount DESC
                LIMIT 10
            """)
            bill_result = self.db.execute(
                bill_query,
                {"company_id": company_id, "start_date": start_date, "end_date": end_date},
            )
            rows = bill_result.fetchall()

        total = sum(Decimal(str(row.amount)) for row in rows) or Decimal("1")

        return [
            {
                "category": row.category or "Uncategorized",
                "amount": Decimal(str(row.amount)),
                "percentage": (Decimal(str(row.amount)) / total * Decimal("100")).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                ),
            }
            for row in rows
        ]

    @staticmethod
    def _safe_percent(numerator: Decimal, denominator: Decimal) -> Decimal:
        """Safe division returning 0 on zero denominator."""
        if denominator == 0:
            return Decimal("0.00")
        return (numerator / denominator).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
